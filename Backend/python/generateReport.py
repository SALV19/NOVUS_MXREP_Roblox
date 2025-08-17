from io import BytesIO
from openpyxl import Workbook
import json
import sys

virtual_workbook = BytesIO()

wb = Workbook()

premises = wb.create_sheet("Premises")
cos = wb.create_sheet("COs")
capacity = wb.create_sheet("Capacity")
bom = wb.create_sheet("BOM")
deprecetion = wb.create_sheet("Deprecetion")
finance = wb.create_sheet("Finance")
expenses = wb.create_sheet("Expenses")
revenue = wb.create_sheet("Revenue")
r_state = wb.create_sheet("R State")
flow = wb.create_sheet("Flow")
balance = wb.create_sheet("Balance")
rates = wb.create_sheet("Rates")
graphs = wb.create_sheet("Graphs")

wb.save(virtual_workbook)

sys.stdout.buffer.write(virtual_workbook.getvalue())
